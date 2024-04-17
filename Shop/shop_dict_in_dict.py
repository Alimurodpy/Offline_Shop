from openpyxl import load_workbook, Workbook
from datetime import datetime, timedelta
import os

line = 49 * '-'
line2 = 49 * '='
line3 = 49 * '^'
line4 = 49 * '*'

class Shop:
    __now_date = datetime.now()
    def __init__(self, file_path) :
        self.file_path = file_path
        self.input_field = dict()
        self.database = dict()
        self.add_database = dict()
        self.buy_database = dict()
        self.report_database = dict()
    
    @classmethod
    def now_date(cls):
        """
               Hozirgi vaqtni qaytaradi   
        """
        cls.__now_date = datetime.now()
        return cls.__now_date
    
    def read_xlsx(self) -> dict:
        """
               Exceldagi barcha ma'lumotlarni lug'at formatida saqlaydi               

        """
        wb = load_workbook(self.file_path)
        global sheet_name 
        sheet_name = wb.sheetnames
        for i in range(1, wb[sheet_name[0]].max_column + 1):
            self.input_field[wb[sheet_name[0]].cell(row = 1, column = i).value] = wb[sheet_name[0]].cell(row = 2, column = i).value
            
        for i in range(len(sheet_name)):
            self.database[sheet_name[i]] = dict()
            for j in range(3, wb[sheet_name[i]].max_row+1):
                self.database[sheet_name[i]][wb[sheet_name[i]].cell(row = j, column = 1).value ] = dict()
                for k in range(2, wb[sheet_name[i]].max_column + 1):
                    self.database[sheet_name[i]][wb[sheet_name[i]].cell(row = j, column = 1).value ][wb[sheet_name[i]].cell(row = 1, column = k).value] = wb[sheet_name[i]].cell(row = j, column = k).value
        for i in range(len(sheet_name)):
            self.add_database[sheet_name[i]] = dict()
        for i in range(len(sheet_name)):
            self.buy_database[sheet_name[i]] = dict()
    def updata_data_excel(self) :
        wb1 = Workbook()
        one_ws = wb1.active
        one_ws.title = sheet_name[0]
        one_ws.append(['name',	'qty',	'price', 'date', 'size', 'color'])
        two_ws = wb1.create_sheet(sheet_name[1])
        two_ws.append(['name',	'qty',	'price', 'date', 'size', 'color'])
        for i in range(2,len(sheet_name)):
            ws = wb1.create_sheet(sheet_name[i])
            ws.append(['name', 'qty', 'price', 'date', 'size', 'color'])
        arr = []
        lines = [[] for _ in range(6)]
        for key in self.database.keys():
            for i, j in self.database[key].items():
                arr.append(str(i))
                for k, d in j.items():
                    arr.append(str(d))
        lines = [arr[i:i+6] for i in range(0, len(arr), 6)]
        s = 1
        for l in lines:
            if s <= len(self.database.get(sheet_name[0]).keys()) :
                s += 1
                one_ws.append(l)
            else :
                two_ws.append(l)            

        wb1.save("Database.xlsx")
        print("\n<<Database.xlsx>> fayli yangilandi.\n")          

    def add_data_excel(self) :
        wb1 = Workbook()
        one_ws = wb1.active
        one_ws.title = sheet_name[0]
        one_ws.append(['name',	'qty',	'price', 'date', 'size', 'color'])
        two_ws = wb1.create_sheet(sheet_name[1])
        two_ws.append(['name',	'qty',	'price', 'date', 'size', 'color'])
        for i in range(2,len(sheet_name)):
            ws = wb1.create_sheet(sheet_name[i])
            ws.append(['name', 'qty', 'price', 'date', 'size', 'color'])
        arr = []
        lines = [[] for _ in range(6)]
        for key in self.add_database.keys():
            for i, j in self.add_database[key].items():
                arr.append(str(i))
                for k, d in j.items():
                    arr.append(str(d))
        lines = [arr[i:i+6] for i in range(0, len(arr), 6)]
        s = 1
        for l in lines:
            if s <= len(self.database.get(sheet_name[0]).keys()) :
                s += 1
                one_ws.append(l)
            else :
                two_ws.append(l)            

        wb1.save("add_data.xlsx")
        print("\n<<add_data.xlsx>> fayli yangilandi.\n")       

    def sell_data_excel(self) :
        wb1 = Workbook()
        one_ws = wb1.active
        one_ws.title = sheet_name[0]
        one_ws.append(['name',	'qty',	'date'])
        two_ws = wb1.create_sheet(sheet_name[1])
        two_ws.append(['name',	'qty',	'date'])
        for i in range(2,len(sheet_name)):
            ws = wb1.create_sheet(sheet_name[i])
            ws.append(['name',	'qty',	'date'])
        arr = []

        for a,v in self.buy_database.items() :
            for k, i in v.items():                                
                for c in range(len(i)):
                    arr.append([k, str(self.buy_database[a][k][c]['qty']), str(self.buy_database[a][k][c]['date'])  ])

        s = 1
        print("arr : ", arr )
        for l in arr:
            if s < len(self.database.get(sheet_name[0]).keys()) :
                s += 1
                one_ws.append(l)
            else :
                two_ws.append(l)              

        wb1.save("sell_data.xlsx")
        print("\n<<sell_data.xlsx>> fayli yangilandi.\n")       


    def write_dict(self, add_miqdor, add_price, add_time, add_size, add_color) -> dict:
        datacha = {
                    "qty" : add_miqdor,
                    "price" : add_price, 
                    "data" : add_time,
                    "size" : add_size,
                    "color" : add_color,             
                } 
        return datacha

    def add_product(self):
        while True :
            for g in range(len(sheet_name)):
                print(f"{g+1} - {sheet_name[g]}") 
            part   = int(input(f"Mahsulotni turini kiriting/Chiqish - 0 : "))
            if part == 0:
                break
            if part <= len(sheet_name):
                part = sheet_name[part-1]
            else :
                print("Qaytadan kiriting!!!")
                continue
            
            add_pruduct   = str(input("Mahsulotni kiriting/Chiqish - 0 : "))
            add_miqdor  = str(input("Miqdorini kiriting : "))
            if add_miqdor.isnumeric() and int(add_miqdor)>0:
                add_miqdor = int(add_miqdor) 
            else:
                print("0 dan katta <son> kiriting")
                continue
            while True:
                add_price   = str(input("Narxini kiriting : "))
                if add_price.isnumeric() and int(add_price)>0 :
                    add_price = int(add_price)
                else:
                    print("0 dan katta son kiriting!!!")
                    continue 
                add_time    = Shop.now_date().strftime("%x %X")
                while True:
                    add_size = str(input("O'lchamini kiriting : "))
                    if add_size.isnumeric() and int(add_size)>0 :
                        add_size = int(add_size)
                    else:
                        print("0 dan katta son kiriting!!!")
                        continue                    
                    add_color = str(input("Rangini kiriting : "))
                    break
                if add_pruduct  in self.database[part].keys():
                    self.database[part][add_pruduct]['qty']+=add_miqdor
                else :
                    self.database[part][add_pruduct] = self.write_dict(add_miqdor, add_price, add_time, add_size, add_color)
                
                if add_pruduct in self.add_database[part].keys():
                    self.add_database[part][add_pruduct]['qty']+=add_miqdor
                else : 
                    self.add_database[part][add_pruduct] = dict()
                    self.add_database[part][add_pruduct] = self.write_dict(add_miqdor, add_price, add_time, add_size, add_color)
                    print("Muvoffaqiyatli qo'shildi!!!") 
                break
                
        self.main()                   

    def sell_product(self):
        sum=0 
        while True :
            for g in range(len(sheet_name)):
                print(f"{g+1} - {sheet_name[g]}") 
            part2   = int(input(f"Mahsulotni turini kiriting/Chiqish - 0 : "))
            if part2 == 0:
                break
            if part2 <= len(sheet_name):
                part2 = sheet_name[part2-1]
            else :
                print("Qaytadan kiriting!!!")
                continue
            while True :
                buy_product = str(input("0)chiqish va chek \nMaxsulotni kiriting : "))
                if buy_product == '0' :
                    print(f"{line3} \n{"🛒 CHEAP_SHOP 🛒".center(len(line3)-2)}")
                    print(f"{"ALIMUROD EKO MARKET".center(len(line3)-2)}")
                    print(f"{"Telefon : +998943672277".center(len(line3)-2)}")
                    x =  Shop.now_date()
                    print(f"{"Data :".rjust(len(line3)//2-3)} {x.strftime("%Y-%m-%d %X ").ljust(len(line3)//2-3)}")                    
                    print(f"{line4}")
                    print(f"{"Mahsulot".ljust(len(line3)//3)}{"Miqdori × narx".ljust(len(line4)//4)}{"Narxi".rjust(19)}")
                    for a,v in self.buy_database.items() :
                        for k, i in v.items():
                            for c in range(len(i)):
                                print(f"{str(k).ljust(len(line3)//3)}{str(self.buy_database[a][k][c]['qty']).ljust(len(line4)//10)} × {str(self.buy_database[a][k][c]['price']).ljust(len(line4)//8)} {str(self.buy_database[a][k][c]['qty'] * self.buy_database[a][k][c]['price']).rjust(18)}")
                                sum+=self.buy_database[a][k][c]['qty'] * self.buy_database[a][k][c]['price']
                    print(f"{line2} \nsh.j.QQS 5% \nJami : {str(sum).rjust(len(line2)-7)}\n{line3}")
                    break                                
                
                
                if self.database[part2].get(buy_product, '1') == '1' :
                    print("Bunday turdagi mahsulot mavjud emas!!!\nQaytadan kiriting.") 
                    continue
                elif self.database[part2][buy_product]["qty"] == 0 :
                    self.database[part2].pop(buy_product)
                    print("Bunday turdagi mahsulot tugagan!!!\nQaytadan kiriting.")
                    continue

                while True :
                    buy_miqdor = str(input("Miqdorini kiriting : "))
                    if buy_miqdor.isalpha():
                        print("Iltimos, son kiriting") 
                        continue
                    elif int(buy_miqdor) <= self.database[part2][buy_product]["qty"]:
                        buy_miqdor = int(buy_miqdor)                        
                        self.database[part2][buy_product]["qty"] -= buy_miqdor
                        if buy_product not in self.buy_database[part2].keys():
                            self.buy_database[part2][buy_product] = []
                        narxi = self.database[part2][buy_product]["price"]
                        narxi += narxi*0.05 
                        x = Shop.now_date()
                        self.buy_database[part2][buy_product].append({"qty" : buy_miqdor, "price" : narxi, "date" : x.strftime("%x %X")})
                        
                        print("Muvoffaqiyatli sotildi!!!")
                        break  
                    else:   
                        print(f"Bizda <<{buy_product}>>dan {self.database[part2][buy_product]["qty"]} miqdorda qolgan!!! \nKamroq miqdorda kiriting") 
                        continue                
        
        self.main()  
    def report_product(self):
        while True :
            part3 = input(f"0)chiqish \n1)Kelgan mahsulotlar \n2)Sotilgan mahsulotlar \n3)Joriy database \n>>>" )
            if part3 == '1':
                for key in self.add_database.keys():
                    print(f"{line2}\n|{key.center(len(line)-2)}|\n{line2}")
                    for i, j in self.add_database[key].items():
                        print(f"|{i.upper().center(len(line)-2)}| \n{line}")
                        for k, d in j.items():
                            print(f"| {str(k).rjust(len(line)//2-3)} : {str(d).ljust(len(line)//2-3)} |")
                        print(line)
            elif part3 == '2':                       
                print(f"{line} \n|{"Mahsulot".rjust(len(line)//4-3)} | {"Miqdor".center((len(line)-10)//2-3)} | {"Sana".center((len(line)-10)//2-4)} |\n{line}")
                for a,v in self.buy_database.items() :
                    for k, i in v.items():                                
                        for c in range(len(i)):                                    
                            print(f"|{str(k).rjust(len(line)//4-4)}  {str(self.buy_database[a][k][c]['qty']).center((len(line)-10)//2-3)}  {str(self.buy_database[a][k][c]['date']).center((len(line)-10)//2-1)} |")
                print(line)
            elif part3 == '3':
                for key in self.database.keys():
                    print(f"{line2}\n|{key.center(len(line)-2)}|\n{line2}")
                    for i, j in self.database[key].items():
                        print(f"|{i.upper().center(len(line)-2)}| \n{line}")
                        for k, d in j.items():
                            print(f"| {str(k).rjust(len(line)//2-3)} : {str(d).ljust(len(line)//2-3)} |")
                        print(line)
            else : 
                break
        self.main()

    def main(self):
        """
            Mahsulot qo'shish, sotish, hisobot chiqarishni boshqaruvchi funksiya
        """
        
        sorov = input(f"\n{40*'#'}\nDasturni to'xtatish uchun  0 : \nMahsulot qo'shish uchun    1 : \nMahsulot sotib olish uchun 2 : \nHisobot chiqarish uchun    3 :\n>>>")
        if sorov == '1':
            self.add_product()
        elif sorov == '2':
            self.sell_product()
        elif sorov == '3':
            self.report_product()
        elif sorov == '0':
            self.add_data_excel()
            self.sell_data_excel()
            self.updata_data_excel()
            quit()
        else:
            print("siz xato buyruq kiritdingiz!!!!!")
            self.main()

# working main
# shop1 = Shop('data_shop.xlsx')
# shop1.read_xlsx()
# shop1.main()        



